"""Excel 批量导入联系人。"""

import re
from typing import Any

from openpyxl import load_workbook

from models import add_contact

# 电话号码正则：以 1 开头的 11 位数字，或带 +86/区号的号码
_PHONE_PATTERN = re.compile(r"^[+\d][\d\s\-()]{6,20}$")


def _is_phone(value: str) -> bool:
    """判断字符串是否像电话号码。"""
    cleaned = re.sub(r"[\s\-()]", "", value)
    return cleaned.isdigit() and len(cleaned) >= 7


def _clean_phone(value: str) -> str:
    """清理电话号码格式：去除小数点、空格、横线等。"""
    phone = str(value).strip()
    if phone.endswith(".0"):
        phone = phone[:-2]
    return phone


def import_excel(file_path: str) -> dict[str, Any]:
    """从 Excel 文件导入联系人。

    自动识别策略（不依赖表头名称）：
    1. 先按表头关键词匹配「姓名」「电话」「分组」列
    2. 找不到则扫描所有列内容，按数据特征自动识别
    3. 纯号码列表（无表头）也能自动导入

    Args:
        file_path: Excel 文件路径

    Returns:
        {"success": 成功数, "skipped": 跳过数, "errors": 错误列表}
    """
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    # 读取前 10 行数据用于分析
    all_rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
    if not all_rows:
        wb.close()
        return {"success": 0, "skipped": 0, "errors": ["文件为空"]}

    # 第一行作为候选表头
    header_row = [str(c).strip() if c else "" for c in all_rows[0]]

    # 第一步：按表头关键词匹配
    name_col = _find_column(header_row, ["姓名", "name", "联系人", "称呼"])
    phone_col = _find_column(header_row, ["电话", "phone", "电话号码", "号码", "手机", "联系方式"])
    group_col = _find_column(header_row, ["分组", "group", "类别", "备注"])

    data_start = 2  # 默认跳过表头

    # 第二步：表头没匹配到，按列内容自动识别
    if phone_col is None:
        phone_col, name_col, has_header = _auto_detect_columns(all_rows)
        if has_header is False:
            data_start = 1  # 无表头，从第一行开始读
        # 如果自动检测也没找到姓名列，用电话列代替
        if name_col is None and phone_col is not None:
            name_col = phone_col

    if phone_col is None:
        wb.close()
        return {"success": 0, "skipped": 0,
                "errors": ["未找到电话号码列，请确保 Excel 中至少有一列包含电话号码"]}

    result = {"success": 0, "skipped": 0, "errors": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start), start=data_start):
        phone_val = row[phone_col].value if phone_col < len(row) else None
        if not phone_val:
            result["skipped"] += 1
            continue

        phone = _clean_phone(str(phone_val))
        if not _is_phone(phone):
            result["skipped"] += 1
            continue

        # 姓名：优先取姓名列，没有就用电话号码
        name = ""
        if name_col is not None and name_col < len(row) and row[name_col].value:
            name = str(row[name_col].value).strip()
            if name == phone:  # 姓名和电话一样说明没有姓名列
                name = phone
        else:
            name = phone

        group_name = ""
        if group_col is not None and group_col < len(row) and row[group_col].value:
            group_name = str(row[group_col].value).strip()

        try:
            add_contact(name=name, phone=phone, group_name=group_name)
            result["success"] += 1
        except Exception as e:
            result["errors"].append(f"第{row_idx}行导入失败: {e}")

    wb.close()
    return result


def _find_column(headers: list[str], keywords: list[str]) -> int | None:
    """在表头中查找匹配关键词的列索引。"""
    for i, h in enumerate(headers):
        h_lower = h.lower()
        for kw in keywords:
            if kw in h_lower:
                return i
    return None


def _auto_detect_columns(rows: list[tuple]) -> tuple[int | None, int | None, bool | None]:
    """扫描列内容，自动识别电话号码列和姓名列。

    Returns:
        (phone_col, name_col, has_header)
        has_header: True=有表头, False=无表头, None=无法判断
    """
    if not rows:
        return None, None, None

    num_cols = max(len(r) for r in rows) if rows else 0

    phone_col = None
    name_col = None
    phone_score_max = 0

    for col_idx in range(num_cols):
        col_values = []
        for r in rows:
            val = str(r[col_idx]).strip() if col_idx < len(r) and r[col_idx] else ""
            col_values.append(val)

        # 统计该列中电话号码的数量
        phone_count = sum(1 for v in col_values if _is_phone(v))
        # 统计非空、非电话的文本数量（可能是姓名）
        text_count = sum(1 for v in col_values
                         if v and not _is_phone(v) and not v.replace(".", "").isdigit())

        # 电话列：电话号码占比高
        if phone_count >= 2 and phone_count > phone_score_max:
            phone_score_max = phone_count
            phone_col = col_idx
            # 如果有另一列文本占比高，那是姓名列
            for other_idx in range(num_cols):
                if other_idx == col_idx:
                    continue
                other_vals = [str(r[other_idx]).strip() if other_idx < len(r) and r[other_idx] else ""
                              for r in rows]
                other_text = sum(1 for v in other_vals
                                 if v and not _is_phone(v) and not v.replace(".", "").isdigit())
                if other_text >= 2:
                    name_col = other_idx
                    break

    # 判断是否有表头
    has_header = None
    if phone_col is not None:
        first_val = str(rows[0][phone_col]).strip() if phone_col < len(rows[0]) and rows[0][phone_col] else ""
        if _is_phone(first_val):
            has_header = False  # 第一行就是电话号码，说明无表头
        else:
            has_header = True

    return phone_col, name_col, has_header
